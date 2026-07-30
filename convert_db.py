#!/usr/bin/env python3
"""Convert Microgravity_Database_reduced.xlsx to SI units and add derived columns.

All conversions/derivations are written as Excel formulas so each cell shows
how it was calculated. Formatting (fonts, fills, borders, widths, merges) is
preserved; new columns inherit the style of their left neighbour.
"""
import re
from copy import copy
from decimal import Decimal

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.formula.translate import Translator

SRC = "Microgravity_Database_reduced.xlsx"
DST = "Microgravity_Database_converted.xlsx"

TF = 1740.0      # flame temperature K
TINF = 298.15    # ambient temperature K

wb = openpyxl.load_workbook(SRC)
ws = wb["Sheet1"]

DATA_FIRST, DATA_LAST = 3, 5059

NUM_MM = re.compile(r"(\d+(?:\.\d+)?)\s*mm")


def mm_to_m_str(num_str: str) -> str:
    """'0.025' (mm) -> '0.000025' (m) as a clean decimal string."""
    d = Decimal(num_str) / Decimal(1000)
    s = format(d.normalize(), "f")
    return s


# ---------------------------------------------------------------- capture ---
# 1) merged ranges (row-1 group headers) -> unmerge before inserting columns
old_merges = [str(m) for m in list(ws.merged_cells.ranges)]
for m in old_merges:
    ws.unmerge_cells(m)

# 2) column widths / dimension props
old_dims = {}
for letter, dim in ws.column_dimensions.items():
    old_dims[letter] = dict(width=dim.width, hidden=dim.hidden,
                            bestFit=dim.bestFit, auto_size=dim.auto_size)

# 3) formula cells (to re-translate after the shift)
formula_cells = []  # (row, old_col_idx, formula)
for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
    for cell in row:
        if isinstance(cell.value, str) and cell.value.startswith("="):
            formula_cells.append((cell.row, cell.column, cell.value))

# 4) raw values needed for the new columns (parse dims while still in mm)
rows_info = {}
for r in range(DATA_FIRST, DATA_LAST + 1):
    geom = ws.cell(row=r, column=4).value          # D
    dims = ws.cell(row=r, column=5).value          # E
    if geom is None:
        continue
    nums = NUM_MM.findall(str(dims)) if dims is not None else []
    rows_info[r] = (str(geom).strip().lower(), nums, dims)


# ---------------------------------------------------------- column shifts ---
def shift(old_idx: int) -> int:
    if old_idx <= 5:
        return old_idx
    if old_idx <= 10:            # F..J  fuel block
        return old_idx + 2
    if old_idx <= 13:            # K..M  core block
        return old_idx + 3
    if old_idx <= 21:            # N..U  gas block
        return old_idx + 4
    return old_idx + 8           # V..   pressure onward


# insert: 2 after E, 1 after fuel_alpha, 1 after core_cp, 4 before Pressure
ws.insert_cols(6, 2)    # F,G  -> half_thickness, characteristic_length_m
ws.insert_cols(13, 1)   # M    -> fuel_volumetric_heat_capacity_J_m3K
ws.insert_cols(17, 1)   # Q    -> core_volumetric_heat_capacity_J_m3K
ws.insert_cols(26, 4)   # Z..AC -> Reynolds, Peclet, Prandtl, thin_thick_regime

# re-translate pre-existing formulas (cell and its refs shift by same offset)
for r, old_c, f in formula_cells:
    new_c = shift(old_c)
    old_coord = f"{get_column_letter(old_c)}{r}"
    new_coord = f"{get_column_letter(new_c)}{r}"
    ws.cell(row=r, column=new_c).value = Translator(
        f, origin=old_coord).translate_formula(new_coord)

# ------------------------------------------------------------- new headers --
NEW_COLS = {6: "half_thickness", 7: "characteristic_length_m",
            13: "fuel_volumetric_heat_capacity_J_m3K",
            17: "core_volumetric_heat_capacity_J_m3K",
            26: "Reynolds", 27: "Peclet", 28: "Prandtl",
            29: "thin_thick_regime"}
for c, name in NEW_COLS.items():
    ws.cell(row=2, column=c).value = name

# updated unit labels on converted columns
ws.cell(row=2, column=5).value = re.sub(r"\bmm\b", "m",
                                        str(ws.cell(row=2, column=5).value))
ws.cell(row=2, column=30).value = "Pressure (Pa)"          # AD (was V)
ws.cell(row=2, column=32).value = "Flow Velocity (m/s)"    # AF (was X)
ws.cell(row=2, column=42).value = "FSR (m/s)"              # AP (was AH)

# ----------------------------------------------------- style for new cols ---
STYLE_SRC = {6: 5, 7: 5, 13: 12, 17: 16, 26: 25, 27: 25, 28: 25, 29: 25}
for c, src in STYLE_SRC.items():
    for r in range(1, DATA_LAST + 1):
        tgt = ws.cell(row=r, column=c)
        ref = ws.cell(row=r, column=src)
        tgt.font = copy(ref.font)
        tgt.fill = copy(ref.fill)
        tgt.border = copy(ref.border)
        tgt.alignment = copy(ref.alignment)
        tgt.number_format = "General"

# ------------------------------------------------- in-place unit formulas ---
def num_str(v):
    if isinstance(v, int):
        return str(v)
    return repr(v)

for r in range(DATA_FIRST, DATA_LAST + 1):
    # Pressure kPa -> Pa (AD = 30)
    v = ws.cell(row=r, column=30).value
    if isinstance(v, (int, float)):
        ws.cell(row=r, column=30).value = f"={num_str(v)}*1000"
    # Flow mm/s -> m/s (AF = 32)
    v = ws.cell(row=r, column=32).value
    if isinstance(v, (int, float)):
        ws.cell(row=r, column=32).value = f"={num_str(v)}/1000"
    # FSR mm/s -> m/s (AP = 42); leave non-numeric notes untouched
    v = ws.cell(row=r, column=42).value
    if isinstance(v, (int, float)):
        ws.cell(row=r, column=42).value = f"={num_str(v)}/1000"

# ------------------------------------- dimensions text mm -> m + new cols ---
unparsed = []
for r, (geom, nums, dims) in rows_info.items():
    # rewrite the dimension string in metres
    if dims is not None:
        new_txt, n = NUM_MM.subn(lambda m: mm_to_m_str(m.group(1)) + " m",
                                 str(dims))
        ws.cell(row=r, column=5).value = new_txt

    m = [mm_to_m_str(x) for x in nums]  # metre strings

    half = lc = None
    if geom == "flat" and len(m) == 3:
        half = f"={m[2]}/2"                       # W/2
        lc = f"={m[0]}"                           # length (first value)
    elif geom == "cylindrical" and len(m) == 2:
        half = f"=({m[1]}-{m[0]})/2"              # (outer - inner)/2
        lc = f"={m[1]}"                           # outer diameter
    elif geom == "wire" and len(m) == 2:
        half = f"=({m[1]}/2)/2"                   # radius/2 (D = 2nd value)
        lc = f"={m[1]}"                           # diameter
    elif geom == "spherical" and len(m) == 1:
        half = f"=({m[0]}/2)/2"                   # radius/2
        lc = f"={m[0]}"                           # diameter
    else:
        unparsed.append((r, geom, dims))

    if half:
        ws.cell(row=r, column=6).value = half
        ws.cell(row=r, column=7).value = lc

    # ---- volumetric heat capacities (density * cp) ----
    ws.cell(row=r, column=13).value = (
        f'=IF(OR($H{r}="",$J{r}=""),"",$H{r}*$J{r})')
    ws.cell(row=r, column=17).value = (
        f'=IF(OR($N{r}="",$P{r}=""),"",$N{r}*$P{r})')

    # ---- dimensionless numbers ----
    flow = f"IF(N($AF{r})=0,0.001,ABS($AF{r}))"   # 0.001 m/s when 0/blank
    ws.cell(row=r, column=26).value = (            # Reynolds = |u|*Lc/nu
        f'=IF($Y{r}="","",{flow}*$G{r}/$Y{r})')
    ws.cell(row=r, column=27).value = (            # Peclet = |u|*Lc/alpha
        f'=IF($X{r}="","",{flow}*$G{r}/$X{r})')
    ws.cell(row=r, column=28).value = (            # Prandtl = nu/alpha
        f'=IF(OR($X{r}="",$Y{r}=""),"",$Y{r}/$X{r})')
    ws.cell(row=r, column=29).value = (            # thin/thick regime
        f'=IF(OR($X{r}="",$U{r}=""),"",'
        f'$G{r}/(SQRT($X{r}*$L{r}/($U{r}*{flow}))'
        f'*({TF}-$K{r})/($K{r}-{TINF})))')

print("unparsed dimension rows:", len(unparsed), unparsed[:10])

# --------------------------------------------------------- column widths ----
new_dims = {}
for letter, props in old_dims.items():
    idx = openpyxl.utils.column_index_from_string(letter)
    new_dims[shift(idx)] = props
ws.column_dimensions.clear()
for idx, props in new_dims.items():
    d = ws.column_dimensions[get_column_letter(idx)]
    if props["width"]:
        d.width = props["width"]
    d.hidden = props["hidden"]
for idx, w in {6: 14, 7: 20, 13: 26, 17: 26, 26: 12, 27: 12, 28: 12, 29: 16}.items():
    ws.column_dimensions[get_column_letter(idx)].width = w

# ------------------------------------------------------------- re-merge -----
for rng in ["A1:C1", "D1:Q1", "R1:AI1", "AJ1:AM1", "AN1:AR1"]:
    ws.merge_cells(rng)

wb.save(DST)
print("saved", DST)